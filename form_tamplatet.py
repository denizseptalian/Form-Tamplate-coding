import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
import datetime
import subprocess

TEMPLATE_FILE = "Tampalatepr.xlsx"

# === helper tulis aman ke cell merge ===
def safe_write(ws, cell, value):
    """
    Tulis ke Excel aman meski cell ada di dalam merge.
    Jika cell termasuk merge range, otomatis tulis ke sel kiri-atas.
    """
    col_str = ''.join(filter(str.isalpha, cell))
    row_str = ''.join(filter(str.isdigit, cell))
    col = column_index_from_string(col_str)
    row = int(row_str)

    for rng in ws.merged_cells.ranges:
        if (row, col) in rng.cells:
            # ambil sel kiri-atas merge
            rmin, cmin, _, _ = rng.bounds
            cell = ws.cell(row=rmin, column=cmin).coordinate
            break

    ws[cell].value = value

# === Streamlit App ===
st.title("Form Purchasing Request (PR)")

with st.form("pr_form"):
    st.header("📋 Header PR")
    departemen = st.text_input("Departemen")
    tanggal = st.date_input("Tanggal Pengajuan", value=datetime.date.today())
    jenis_pekerjaan = st.text_input("Jenis Pekerjaan / Unit / Stasiun")
    no_pp = st.text_input("Nomor PP")

    st.header("📦 Detail Barang")
    if "barang" not in st.session_state:
        st.session_state.barang = [
            {"kode": "", "nama": "", "spesifikasi": "", "satuan": "", "qty": 1, "harga": 0, "keterangan": ""}
        ]

    for i, item in enumerate(st.session_state.barang):
        st.subheader(f"Barang {i+1}")
        st.session_state.barang[i]["kode"] = st.text_input(f"Kode Barang {i+1}", value=item["kode"], key=f"kode{i}")
        st.session_state.barang[i]["nama"] = st.text_input(f"Nama Barang {i+1}", value=item["nama"], key=f"nama{i}")
        st.session_state.barang[i]["spesifikasi"] = st.text_area(f"Spesifikasi {i+1}", value=item["spesifikasi"], key=f"spesifikasi{i}")
        st.session_state.barang[i]["satuan"] = st.text_input(f"Satuan {i+1}", value=item["satuan"], key=f"satuan{i}")
        st.session_state.barang[i]["qty"] = st.number_input(f"Kuantitas {i+1}", min_value=1, value=item["qty"], key=f"qty{i}")
        st.session_state.barang[i]["harga"] = st.number_input(f"Estimasi Harga {i+1}", min_value=0, value=item["harga"], key=f"harga{i}")
        st.session_state.barang[i]["keterangan"] = st.text_area(f"Keterangan {i+1}", value=item["keterangan"], key=f"keterangan{i}")
        st.divider()

    tambah_barang = st.form_submit_button("➕ Tambah Barang", type="secondary")

    st.header("💰 Anggaran")
    total_anggaran = st.number_input("Total Anggaran (Rp)", min_value=0, value=0)
    actual_pengeluaran = st.number_input("Actual Pengeluaran (Rp)", min_value=0, value=0)
    permintaan_saat_ini = st.number_input("Permintaan Saat Ini (Rp)", min_value=0, value=0)
    saldo_anggaran = st.number_input("Saldo Anggaran (Rp)", min_value=0, value=0)

    st.header("👨‍💼 Persetujuan")
    diajukan_ktu = st.text_input("Diajukan Oleh (KTU)")
    diajukan_mgr = st.text_input("Diajukan Oleh (Estate Manager)")

    submitted = st.form_submit_button("💾 Simpan & Cetak")

if tambah_barang:
    st.session_state.barang.append(
        {"kode": "", "nama": "", "spesifikasi": "", "satuan": "", "qty": 1, "harga": 0, "keterangan": ""}
    )
    st.rerun()

# === Isi Excel + Export ke PDF (opsional) ===
if submitted:
    wb = load_workbook(TEMPLATE_FILE)
    ws = wb.active

    # --- isi header ---
    safe_write(ws, "B3", f"Departemen: {departemen}")
    safe_write(ws, "B3", f"Tanggal: {tanggal}")
    safe_write(ws, "B3", f"Jenis Pekerjaan: {jenis_pekerjaan}")
    safe_write(ws, "B3", f"No. PP: {no_pp}")

    # --- isi detail barang (mulai baris 11) ---
    row_start = 11
    for i, item in enumerate(st.session_state.barang):
        r = row_start + i
        safe_write(ws, f"B{r}", item["kode"])
        safe_write(ws, f"C{r}", item["nama"])
        safe_write(ws, f"D{r}", item["spesifikasi"])
        safe_write(ws, f"E{r}", item["satuan"])
        safe_write(ws, f"F{r}", item["qty"])
        safe_write(ws, f"G{r}", item["harga"])
        safe_write(ws, f"H{r}", item["qty"] * item["harga"])
        safe_write(ws, f"I{r}", item["keterangan"])

    # --- isi anggaran (contoh baris 25–27) ---
    safe_write(ws, "B25", f"Total Anggaran: Rp {total_anggaran:,}")
    safe_write(ws, "B27", f"Actual: Rp {actual_pengeluaran:,} | Permintaan: Rp {permintaan_saat_ini:,} | Saldo: Rp {saldo_anggaran:,}")

    # --- persetujuan ---
    safe_write(ws, "B31", diajukan_ktu)  # atas tanda tangan KTU
    safe_write(ws, "F31", diajukan_mgr)  # atas tanda tangan Estate Manager

    # simpan Excel
    output_xlsx = "PR_Output.xlsx"
    wb.save(output_xlsx)

    # --- convert Excel ke PDF jika ada LibreOffice ---
    output_pdf = "PR_Output.pdf"
    try:
        subprocess.run(
            ["libreoffice", "--headless", "--convert-to", "pdf", output_xlsx, "--outdir", "."],
            check=True
        )
        with open(output_pdf, "rb") as f:
            st.success("✅ PR berhasil dibuat dalam format PDF")
            st.download_button("⬇️ Download PR PDF", f, file_name=output_pdf)
    except Exception:
        st.warning("📂 PR berhasil dibuat dalam format Excel. (PDF otomatis butuh LibreOffice/Excel).")
        with open(output_xlsx, "rb") as f:
            st.download_button("⬇️ Download PR Excel", f, file_name=output_xlsx)
